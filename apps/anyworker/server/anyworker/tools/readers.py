"""File-reading tools for local RAG — txt, pdf, xlsx, csv, md."""

from __future__ import annotations

import csv
import io
import logging
from pathlib import Path
from typing import Any

log = logging.getLogger(__name__)


def _resolve_path(workspace: str, rel_path: str) -> Path:
    """Resolve a workspace-relative path safely (no escape of workspace)."""
    base = Path(workspace).resolve()
    target = (base / rel_path).resolve()
    if not target.is_relative_to(base):
        raise ValueError(f"Path escapes workspace: {rel_path}")
    return target


def _guess_extension(path: Path) -> str:
    return path.suffix.lower()


def read_text(workspace: str, rel_path: str) -> dict[str, Any]:
    """Read a plain-text file and return its content."""
    try:
        path = _resolve_path(workspace, rel_path)
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}
    if not path.exists():
        return {"ok": False, "error": f"file not found: {rel_path}"}
    try:
        content = path.read_text(encoding="utf-8", errors="replace")
        return {
            "ok": True,
            "path": str(path),
            "content": content,
            "size": path.stat().st_size,
        }
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


def read_pdf(workspace: str, rel_path: str) -> dict[str, Any]:
    """Extract text content from a PDF file."""
    try:
        import pymupdf  # type: ignore
    except ImportError:
        return {"ok": False, "error": "pymupdf not installed — add pymupdf to dependencies"}

    path = _resolve_path(workspace, rel_path)
    if not path.exists():
        return {"ok": False, "error": f"file not found: {rel_path}"}
    if path.suffix.lower() != ".pdf":
        return {"ok": False, "error": f"not a PDF file: {rel_path}"}

    try:
        doc = pymupdf.open(str(path))
        pages_text: list[str] = []
        for i, page in enumerate(doc):
            text = page.get_text()
            pages_text.append(f"--- page {i + 1} ---\n{text}")
        doc.close()
        return {
            "ok": True,
            "path": str(path),
            "pages": len(pages_text),
            "content": "\n".join(pages_text),
        }
    except Exception as exc:
        log.exception("PDF read failed for %s", path)
        return {"ok": False, "error": str(exc)}


def read_xlsx(workspace: str, rel_path: str) -> dict[str, Any]:
    """Extract tabular data from an XLSX file."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return {"ok": False, "error": "openpyxl not installed — add openpyxl to dependencies"}

    path = _resolve_path(workspace, rel_path)
    if not path.exists():
        return {"ok": False, "error": f"file not found: {rel_path}"}
    if path.suffix.lower() not in {".xlsx", ".xls"}:
        return {"ok": False, "error": f"not an Excel file: {rel_path}"}

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheets: dict[str, Any] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str | None]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            sheets[sheet_name] = {"headers": rows[0] if rows else [], "rows": rows[1:] if len(rows) > 1 else []}
        wb.close()
        return {"ok": True, "path": str(path), "sheets": sheets}
    except Exception as exc:
        log.exception("XLSX read failed for %s", path)
        return {"ok": False, "error": str(exc)}


def read_csv(workspace: str, rel_path: str) -> dict[str, Any]:
    """Parse a CSV file and return rows with headers."""
    path = _resolve_path(workspace, rel_path)
    if not path.exists():
        return {"ok": False, "error": f"file not found: {rel_path}"}
    if path.suffix.lower() != ".csv":
        return {"ok": False, "error": f"not a CSV file: {rel_path}"}

    try:
        content = path.read_text(encoding="utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(content))
        headers = reader.fieldnames or []
        rows = [dict(row) for row in reader]
        return {"ok": True, "path": str(path), "headers": headers, "row_count": len(rows), "rows": rows}
    except Exception as exc:
        log.exception("CSV read failed for %s", path)
        return {"ok": False, "error": str(exc)}


def read_file(workspace: str, rel_path: str) -> dict[str, Any]:
    """Read any supported file and return its content.

    Dispatches to the appropriate reader based on file extension.
    Supported: .txt, .md, .csv, .pdf, .xlsx, .xls
    """
    path = _resolve_path(workspace, rel_path)
    if not path.exists():
        return {"ok": False, "error": f"file not found: {rel_path}"}

    ext = path.suffix.lower()
    if ext in {".txt", ".md", ".log", ".json", ".yaml", ".yml", ".toml", ".cfg", ".ini", ".xml", ".html", ".css", ".js", ".ts", ".py"}:
        result = read_text(workspace, rel_path)
    elif ext == ".pdf":
        result = read_pdf(workspace, rel_path)
    elif ext in {".xlsx", ".xls"}:
        result = read_xlsx(workspace, rel_path)
    elif ext == ".csv":
        result = read_csv(workspace, rel_path)
    else:
        # Fallback: try reading as text
        result = read_text(workspace, rel_path)

    return result


def list_supported_extensions() -> list[str]:
    """Return the list of file extensions supported by read_file."""
    return [".txt", ".md", ".log", ".json", ".yaml", ".yml", ".toml", ".cfg", ".ini", ".xml", ".html", ".css", ".js", ".ts", ".py", ".csv", ".pdf", ".xlsx", ".xls"]
